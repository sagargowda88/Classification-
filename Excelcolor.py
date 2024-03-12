import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# Load the merged DataFrame
merged_data = pd.read_csv('merged_test_and_rows_not_in_test.csv')

# Create a new Excel workbook and add a worksheet
wb = Workbook()
ws = wb.active

# Write the DataFrame to the worksheet
for r_idx, row in enumerate(merged_data.itertuples(), start=1):
    for c_idx, value in enumerate(row[1:], start=1):
        ws.cell(row=r_idx, column=c_idx, value=value)

# Apply color to the 'Predicted_Label' column based on wrong predictions
for r_idx, row in enumerate(merged_data.itertuples(), start=1):
    if row.Predicted_Label != row.target_column:  # Check for wrong predictions
        cell = ws.cell(row=r_idx, column=merged_data.columns.get_loc('Predicted_Label') + 1)
        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

# Save the workbook
wb.save('merged_test_and_rows_not_in_test_styled.xlsx')

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# Load the CSV file
data = pd.read_csv('your_file.csv')

# Create a new Excel workbook and add a worksheet
wb = Workbook()
ws = wb.active

# Write the data to the worksheet
for r_idx, row in enumerate(data.itertuples(), start=1):
    for c_idx, value in enumerate(row[1:], start=1):
        ws.cell(row=r_idx, column=c_idx, value=value)

# Apply style to highlight cells where values in two columns don't match
for r_idx, row in enumerate(data.itertuples(), start=1):
    if row.column1 != row.column2:  # Adjust column1 and column2 based on your CSV
        cell1 = ws.cell(row=r_idx, column=data.columns.get_loc('column1') + 1)
        cell2 = ws.cell(row=r_idx, column=data.columns.get_loc('column2') + 1)
        cell1.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        cell2.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

# Save the workbook
wb.save('highlighted_data.xlsx')
